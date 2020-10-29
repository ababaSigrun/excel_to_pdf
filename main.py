import os
import glob
import openpyxl
import sys
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase import pdfmetrics
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
# pip install openpyxlをやってExcelを開ける環境を作ること。

#  出力先の中身を空にする。
def delPDF ():
    print('ファイル消去メソッド呼び出し')
    # for file in glob.glob('./OutputPDF/*.pdf'):
        # os.remove(file)


# ファイルを開きExcel内部でIDが書いてあるセルを特定する。
def getPassIDText():
    print('IDセル特定呼出')
    f = open('./ImportPasswordText/passAddressList.txt')
    lines = f.readlines()
    count = 0
    splitWord = '=' # 区切り文字を定義
    rowLine = ""
    colLine = ""
    for line in lines:
        count = count + 1
        print(line)
        if line.find("Column") !=-1 : 
            print('Column行' + line)
            colLine = line
        elif line.find("Row")  != -1: 
            print('Row行' + line)
            rowLine = line
        else:
            print('関係ない行が含まれています   : ' + str(count) + '行目')

    return getSplit(colLine,splitWord),getSplit(rowLine,splitWord)

# 文字列を分割し、改ページがあったら削除する
def getSplit(word,splitWord):
    retWord = word.split(splitWord)[1].replace('\n','')
    return retWord


# パスリストを取得する
def getPassList():
    f = open('./getPassIDText/passList.txt')
    lines = f.readlines()
    passList = []
    for line in lines:
        passList.append(line.replace('\n',''))
    return passList

# 元のExcelのファイル名を取得する。
def getTagetFileNameList():
    retList = []
    f_path = 'ImportExcel'
    for f in glob.glob('ImportExcel/*.xlsx'):
        retList.append(os.path.split(f)[1])
    return retList

#Excelファイルを開く。
def getExcel(fileName,targetAddress,sheetName,passList) :
    wb=openpyxl.load_workbook('ImportExcel/' + fileName)
    sheet = wb.get_sheet_by_name(sheetName)
    # 対象のセルのアドレスを取得
    targetCel = sheet[targetAddress].value
    for passid in passList:
        if targetCel ==  passid.split("=")[0] :
            #一致するものがあった場合に以下処理を行う。
            print(fileName)
            # コピーを作る。
            ws = wb.active         
            # OutputPDF
            pdf_gen_proc(fileName.replace('xlsx','pdf'),ws)
            # 対象のファイルを探しパスを設定
            # Dirの移動
        else :
            print(targetCel + "  : " + passid)



# https://qiita.com/ekzemplaro/items/125a578167d41b3540ef 参照
def pdf_gen_proc(file_pdf,ws):
    doc = SimpleDocTemplate(file_pdf, pagesize=A4)
    fontname_g = "HeiseiKakuGo-W5"
    pdfmetrics.registerFont(UnicodeCIDFont(fontname_g))
    elements = []
#
    data = []
    for row in ws.rows:
        unit_aa = []
        print(row[0].value)
        for col in row:
            unit_aa.append(col.value)
        data.append(unit_aa)
    tt=Table(data)
    tt.setStyle(TableStyle([('BACKGROUND',(1,1),(-2,-2),colors.cyan),
        ('TEXTCOLOR',(0,0),(1,-1),colors.red),
        ('FONT', (0, 0), (-1, -1), "HeiseiKakuGo-W5", 20),
        ('GRID', (0, 0), (ws.max_column, ws.max_row), 0.25, colors.black),]))
    elements.append(tt)
    # print('ws.max_column : ' + str(ws.max_column) + ' ws.max_row : '+ str(ws.max_row))
#
    doc.build(elements)


# メインメソッド
print('スタート')
# 出力先の削除を行う。
delPDF()

# 指定座標を取得する。
col,row = getPassIDText()
# 座標生成
targetAddress = col + str(row)

# パスリストを取得する。
passList = getPassList()

# 対象のファイル名を取得する。
tagetFileNameList = getTagetFileNameList()




# 対象のファイルを開きpdfのコピーを作る
for name in tagetFileNameList :
    getExcel(name,targetAddress,'Sheet1',passList)

# https://fastclassinfo.com/entry/python_excel_pdf/